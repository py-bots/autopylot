from auto_pylot.CrashHandler import report_error


def excel_create_excel_file_in_given_folder(fullPathToTheFolder="", excelFileName="", sheet_name="Sheet1"):
    """
    Creates an excel file in the desired folder with desired filename
    Internally this uses folder_create() method to create folders if the folder/s does not exist.
    Parameters:
        fullPathToTheFolder (str) : Complete path to the folder with double slashes.
        excelFileName       (str) : File Name of the excel to be created (.xlsx extension will be added automatically.
        sheet_name           (str) : By default it will be "Sheet1".

    Returns:
        returns boolean TRUE if the excel file is created
    """
    # import section
    from openpyxl import Workbook
    from auto_pylot.Engine import folder_create
    import os
    from pathlib import Path
    status = False
    data = None
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not fullPathToTheFolder:
            raise Exception("Please provide the full path to the folder")

        if not excelFileName:
            raise Exception("Please provide the file name")

        folder_create(fullPathToTheFolder)

        if ".xlsx" in excelFileName:
            excel_path = os.path.join(fullPathToTheFolder, excelFileName)
        else:
            excel_path = os.path.join(
                fullPathToTheFolder, excelFileName + ".xlsx")

        excel_path = Path(excel_path)

        wb.save(filename=excel_path)

    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_set_single_cell(excel_path="", sheet_name="Sheet1", header=0, columnName="", cellNumber=0, setText=""):

    # Description:
    """
    Writes the given text to the desired column/cell number for the given excel file
    """

    # import section
    from auto_pylot.Engine import append_df_to_excel
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Please provide the excel path")
        if not sheet_name:
            raise Exception("Please provide the sheet name")
        if not isinstance(header, int):
            raise Exception("Please provide the header number")
        if not columnName:
            raise Exception("Please provide the column name")

        if not setText:
            raise Exception("Please provide the text to be set")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')

        df.at[cellNumber, columnName] = setText
        append_df_to_excel(excel_path, df, index=False, startrow=0)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_get_single_cell(excel_path="", sheet_name="Sheet1", header=0, columnName="", cellNumber=0):

    # Description:
    """
    Gets the text from the desired column/cell number of the given excel file
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Please provide the excel path")
        if not sheet_name:
            raise Exception("Please provide the sheet name")
        if not isinstance(header, int):
            raise Exception("Please provide the header number")
        if not columnName:
            raise Exception("Please provide the column name")

        if not isinstance(columnName, list):
            columnName = [columnName]

        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols={
            columnName[0]}, engine='openpyxl')
        cellValue = df.at[cellNumber, columnName[0]]
        data = cellValue

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_get_all_header_columns(excel_path="", sheet_name="Sheet1", header=0):

    # Description:
    """
    Gives you all column header names of the given excel sheet.
    """

    # import section
    import pandas as pd
    # Response section
    status = False
    data = None

    try:
        if not excel_path:
            raise Exception("Please provide the excel path")
        if not sheet_name:
            raise Exception("Please provide the sheet name")
        if not isinstance(header, int):
            raise Exception("Please provide the header number")

        col_lst = pd.read_excel(excel_path, sheet_name=sheet_name, header=header,
                                nrows=1, dtype=str, engine='openpyxl').columns.tolist()
        data = col_lst

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_get_all_sheet_names(excelFilePath=""):
    # Description:
    """
    Gives you all names of the sheets in the given excel sheet.

    Parameters:
        excelFilePath  (str) : Full path to the excel file with slashes.

    returns : 
        all the names of the excelsheets as a LIST.
    """

    # import section
    from openpyxl import load_workbook

    # Response section
    status = False
    data = None

    try:
        if not excelFilePath:
            raise Exception("Please provide the excel path")

        wb = load_workbook(excelFilePath)
        data = wb.sheetnames

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_get_row_column_count(excel_path="", sheet_name="Sheet1", header=0):

    # Description:
    """
    Gets the row and coloumn count of the provided excel sheet.

    Parameters:
        excel_path  (str) : Full path to the excel file with slashes.
        sheet_name           (str) : by default it is Sheet1.

    Returns:
        row (int) : number of rows
        col (int) : number of coloumns
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Excel path is empty.")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')
        row, col = df.shape
        row = row + 1
        data = row, col

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_copy_range_from_sheet(excel_path="", sheet_name='Sheet1', startCol=0, startRow=0, endCol=0, endRow=0):

    # Description:
    """
    Copies the specific range from the provided excel sheet and returns copied data as a list
    Parameters:
        excel_path :"Full path of the excel file with double slashes"
        sheet_name     :"Source sheet name from where contents are to be copied"
        startCol          :"Starting column number (index starts from 1) from where copying starts"
        startRow          :"Starting row number (index starts from 1) from where copying starts"
        endCol            :"Ending column number ex:4 upto where cells to be copied"
        endRow            :"Ending column number ex:5 upto where cells to be copied"

    Returns:
    rangeSelected        : the copied range data
    """

    # import section
    from openpyxl import load_workbook

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Excel path is empty.")

        if startCol == 0 and startRow == 0 and endCol == 0 and endRow == 0:
            raise Exception("Please provide the range to be copied.")

        startRow = int(startRow)
        startCol = int(startCol)
        endRow = int(endRow)
        endCol = int(endCol)

        from_wb = load_workbook(filename=excel_path)
        try:
            fromSheet = from_wb[sheet_name]
        except:
            fromSheet = from_wb.worksheets[0]
        rangeSelected = []

        if endRow < startRow:
            endRow = startRow

        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol+1, 1):
                rowSelected.append(
                    fromSheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        data = rangeSelected

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_copy_paste_range_from_to_sheet(excel_path="", sheet_name='Sheet1', startCol=0, startRow=0, endCol=0, endRow=0, copiedData=""):

    # Description:
    """
    Pastes the copied data in specific range of the given excel sheet.
    """

    # import section
    from auto_pylot.Engine import excel_create_excel_file_in_given_folder
    from openpyxl import load_workbook

    # Response section
    status = False
    data = None

    try:

        try:
            if not copiedData:
                raise Exception(
                    "Copied data is empty. First copy the data using Copy Range From Sheet function.")

            if not excel_path:
                raise Exception("Excel path is empty.")

            if startCol == 0 and startRow == 0 and endCol == 0 and endRow == 0:
                raise Exception("Please provide the range to be copied.")

            startRow = int(startRow)
            startCol = int(startCol)
            endRow = int(endRow)
            endCol = int(endCol)

            to_wb = load_workbook(filename=excel_path)
            toSheet = to_wb[sheet_name]

        except:
            try:
                excel_create_excel_file_in_given_folder((str(excel_path[:(str(excel_path).rindex("\\"))])), (str(
                    excel_path[str(excel_path).rindex("\\")+1:excel_path.find(".")])), sheet_name)
            except:
                excel_create_excel_file_in_given_folder((str(excel_path[:(str(excel_path).rindex(
                    "/"))])), (str(excel_path[str(excel_path).rindex("/")+1:excel_path.find(".")])), sheet_name)

            to_wb = load_workbook(filename=excel_path)

            toSheet = to_wb[sheet_name]

        if endRow < startRow:
            endRow = startRow

        countRow = 0
        for i in range(startRow, endRow+1, 1):
            countCol = 0
            for j in range(startCol, endCol+1, 1):
                toSheet.cell(
                    row=i, column=j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        to_wb.save(excel_path)
        data = countRow-1

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_split_by_column(excel_path="", sheet_name='Sheet1', header=0, columnName=""):

    # Description:
    """
    Splits the excel file by Column Name
    """

    # import section
    import pandas as pd
    from pathlib import Path
    import os
    import tempfile

    # Response section
    status = False
    data = None

    try:
        temp_current_working_dir = tempfile.mkdtemp(
            prefix="cloint_", suffix="_fusion")
        temp_current_working_dir = Path(temp_current_working_dir)
        output_folder_path = Path(os.path.join(
            temp_current_working_dir, "Output"))

        if not excel_path:
            raise Exception("Excel path is empty.")

        if not columnName:
            raise Exception("Please provide the column name to be split.")

        data_df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, dtype=str, engine='openpyxl')

        grouped_df = data_df.groupby(columnName)

        for data in grouped_df:
            file_path = os.path.join(
                output_folder_path, str(data[0]) + ".xlsx")
            file_path = Path(file_path)
            grouped_df.get_group(data[0]).to_excel(
                file_path, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_merge_all_files(input_folder_path="", output_folder_path=""):

    # Description:
    """
    Merges all the excel files in the given folder
    """

    # import section
    import os
    import pathlib as Path
    import pandas as pd
    import datetime

    # Response section
    status = False
    data = None

    try:

        if not input_folder_path:
            raise Exception("Input folder path is empty.")

        if not output_folder_path:
            raise Exception("Output folder path is empty.")

        filelist = [f for f in os.listdir(
            input_folder_path) if f.endswith(".xlsx")]
        all_excel_file_lst = []
        for file1 in filelist:
            file_path = os.path.join(input_folder_path, file1)
            file_path = Path(file_path)

            all_excel_file = pd.read_excel(
                file_path, dtype=str, engine='openpyxl')
            all_excel_file_lst.append(all_excel_file)

        appended_df = pd.concat(all_excel_file_lst)
        time_stamp_now = datetime.datetime.now().strftime("%m-%d-%Y")
        final_path = os.path.join(
            output_folder_path, "Final-" + time_stamp_now + ".xlsx")
        final_path = Path(final_path)
        appended_df.to_excel(final_path, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_drop_columns(excel_path="", sheet_name='Sheet1', header=0, columnsToBeDropped=""):

    # Description:
    """
    Drops the desired column from the given excel file
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Excel path is empty.")

        if not columnsToBeDropped:
            raise Exception(
                "Please provide the column name to be dropped.")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')

        if isinstance(columnsToBeDropped, list):
            df.drop(columnsToBeDropped, axis=1, inplace=True)
        else:
            df.drop([columnsToBeDropped], axis=1, inplace=True)

        with pd.ExcelWriter(excel_path) as writer:  # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_clear_sheet(excel_path="", sheet_name="Sheet1", header=0):

    # Description:
    """
    Clears the contents of given excel files keeping header row intact
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Excel path is empty.")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')
        df = df.head(0)

        with pd.ExcelWriter(excel_path) as writer:  # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_remove_duplicates(excel_path="", sheet_name="Sheet1", header=0, columnName="", saveResultsInSameExcel=True, which_one_to_keep="first"):

    # Description:
    """
    Drops the duplicates from the desired Column of the given excel file
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:

        if not excel_path:
            raise Exception("Excel path is empty.")

        if not columnName:
            raise Exception("Please provide the column name.")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')

        count = 0
        if saveResultsInSameExcel:
            df.drop_duplicates(subset=columnName,
                               keep=which_one_to_keep, inplace=True)
            with pd.ExcelWriter(excel_path) as writer:  # pylint: disable=abstract-class-instantiated
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            count = df.shape[0]
        else:
            df1 = df.drop_duplicates(
                subset=columnName, keep=which_one_to_keep, inplace=False)
            excel_path = str(excel_path).replace(".", "_DupDropped.")
            with pd.ExcelWriter(excel_path) as writer:  # pylint: disable=abstract-class-instantiated
                df1.to_excel(
                    writer, sheet_name=sheet_name, index=False)
            count = df1.shape[0]

        data = count

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_describe_data(excel_path="", sheet_name='Sheet1', header=0):

    # Description:
    """
    Describe statistical data for the given excel
    """

    # import section
    import pandas as pd
    # Response section
    status = False
    data = None

    try:
        if not excel_path:
            raise Exception("Excel path is empty.")

        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')

        #user_option_lst = ['Numerical','String','Both']

        #user_choice = gui_get_dropdownlist_values_from_user("list of datatypes",user_option_lst)

        # if user_choice == 'Numerical':
        #    return df.describe(include = [np.number])
        # elif user_choice == 'String':
        #    return df.describe(include = ['O'])
        # else:
        #    return df.describe(include='all')
        data = df.describe()

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_if_value_exists(excel_path="", sheet_name='Sheet1', header=0, usecols="", value=""):

    # Description:
    """
    Check if a given value exists in given excel. Returns True / False
    """

    # import section
    import pandas as pd

    # Response section
    status = False
    data = None

    try:
        if not excel_path:
            raise Exception("Please provide the excel path")
        if not sheet_name:
            raise Exception("Please provide the sheet name")
        if not isinstance(header, int):
            raise Exception("Please provide the header number")

        if not value:
            raise Exception("Please provide the value to be searched")

        if usecols:
            df = pd.read_excel(excel_path, sheet_name=sheet_name,
                               header=header, usecols=usecols, engine='openpyxl')
        else:
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')

        if value in df.values:
            df = ''
            status = True
        else:
            df = ''
            status = False

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def excel_to_colored_html(formatted_excel_path=""):

    # Description:
    """
    Converts given Excel to HTML preserving the Excel format and saves in same folder as .html
    """

    # import section

    # Response section
    status = False
    data = None

    try:
        from xlsx2html import xlsx2html

        if not formatted_excel_path:
            raise Exception("Please provide the excel path")

        formatted_html_path = str(
            formatted_excel_path).replace(".xlsx", ".html")
        xlsx2html(formatted_excel_path, formatted_html_path)
        data = formatted_html_path

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    else:
        status = True
    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def isNaN(value):

    # Description:
    """
    Returns TRUE if a given value is NaN False otherwise
    """

    # import section

    # Response section
    status = False
    data = None

    try:
        if not value:
            raise Exception(
                "Value is empty. Please give a value to check.")
        import math
        status = math.isnan(float(value))

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as e:
        report_error(e)

    finally:
        if status is True and data is not None:
            return [status, data]
        return [status]


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                       truncate_sheet=False, resizeColumns=True, na_rep='NA', **to_excel_kwargs):

    try:
        from string import ascii_uppercase
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        from openpyxl import load_workbook
        import pandas as pd

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        try:
            f = open(filename)
            # Do something with the file
        except IOError:
            # print("File not accessible")
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            wb.save(filename)

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            # startrow = -1
            startrow = 0

        if startcol is None:
            startcol = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow,
                    startcol=startcol, na_rep=na_rep, **to_excel_kwargs)

        if resizeColumns:

            ws = writer.book[sheet_name]

            def auto_format_cell_width(ws):
                for letter in range(1, ws.max_column):
                    maximum_value = 0
                    for cell in ws[get_column_letter(letter)]:
                        val_to_check = len(str(cell.value))
                        if val_to_check > maximum_value:
                            maximum_value = val_to_check
                    ws.column_dimensions[get_column_letter(
                        letter)].width = maximum_value + 2

            auto_format_cell_width(ws)

        # save the workbook
        writer.save()
    except Exception as ex:
        report_error(ex)
